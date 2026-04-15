"""Financial Planner - main entry point.
Loads data, renders sidebar, routes to per-tab modules.
"""
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import csv
import json
# -- Config -------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "Main Financial Plan.xlsx"
POSITION_DIR = BASE_DIR / "Position Files"
st.set_page_config(page_title="Financial Planner", layout="wide", page_icon="\U0001f4b0")
from tabs.shared import (
    MONEY_MARKET_TICKERS, MUTUAL_FUND_TICKERS, KNOWN_YIELDS,
    EXPENSE_FIELDS, INVEST_FIELDS, INCOME_FIELDS, NUM_MONTHS,
    ROW_LABELS, INVEST_TO_PORTFOLIO, ALL_BUDGET_FIELDS,
    CAR_LOAN_FILE, RETIREMENT_FILE, PROJECTION_FILE, CREDIT_CARD_BAL_FILE, LAST_SAVED_FILE,
    save_all, load_all, calc_annual_taxes, compute_paycheck_schedule, compute_month_totals,
)
# -- Parse Position CSV -------------------------------------------------------
def parse_dollar(s):
    if s is None:
        return 0.0
    s = str(s).replace("$", "").replace(",", "").replace("(", "-").replace(")", "").strip()
    if s in ("", "N/A", "<empty>"):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0
def parse_position_csv(filepath):
    content = Path(filepath).read_text(encoding="utf-8-sig")
    lines = content.splitlines()
    equities, others = [], []
    section = None
    header = None
    cash_amount = 0.0
    for line in lines:
        row = next(csv.reader([line]))
        if not row or not row[0].strip():
            continue
        first = row[0].strip()
        if first.startswith("Position Statement"):
            continue
        if first == "Equities and Equity Options":
            section = "equities"; continue
        if first == "Others":
            section = "others"; continue
        if first == "Instrument" and section:
            header = [c.strip() for c in row]; continue
        if first.startswith("Cash & Sweep"):
            for part in row:
                v = parse_dollar(part)
                if v != 0:
                    cash_amount = v
            section = None; continue
        skips = ("OVERALL", "BP ", "OVERNIGHT", "Subtotal", "Overall", "P/L", "Available")
        if any(first.startswith(s) for s in skips):
            continue
        if header and section:
            if len(first) <= 6 and first.replace(".", "").isalpha():
                record = {}
                for i, h in enumerate(header):
                    record[h] = row[i].strip() if i < len(row) else ""
                (equities if section == "equities" else others).append(record)
    def build_df(records):
        if not records:
            return pd.DataFrame()
        df = pd.DataFrame(records)
        if "Qty" in df.columns:
            df["Qty"] = df["Qty"].apply(lambda x: parse_dollar(str(x).replace("+", "")))
        for col in ["Trade Price", "Mark", "Net Liq", "Total Cost", "P/L Open", "P/L Day", "Mrk Chng"]:
            if col in df.columns:
                df[col] = df[col].apply(parse_dollar)
        return df
    return build_df(equities), build_df(others), cash_amount, content
@st.cache_data
def load_latest_positions():
    POSITION_DIR.mkdir(parents=True, exist_ok=True)
    csvs = list(POSITION_DIR.glob("*.csv"))
    for f in BASE_DIR.glob("*PositionStatement*.csv"):
        if f not in csvs:
            csvs.append(f)
    if not csvs:
        return pd.DataFrame(), pd.DataFrame(), 0.0, ""
    csvs = sorted(csvs, key=lambda f: f.name)
    return parse_position_csv(csvs[-1])
# -- Load Excel defaults ------------------------------------------------------
@st.cache_data
def load_excel_defaults():
    _defaults_only = {
        "net_pay": 4430.0, "rent": 1100.0, "car_insurance_gym": 512.0,
        "food": 700.0, "gas": 150.0, "car_payment": 1000.0,
        "brokerage_main": 250.0, "btc": 300.0, "savings": 1000.0,
        "income_brokerage": 0.0, "bonus": 0.0,
    }
    if not EXCEL_FILE.exists():
        return _defaults_only, {"Car Loan": 0.0}, 0.06, 0.0275, {}, {}
    try:
        import openpyxl
    except ImportError:
        # Keep app running on environments missing Excel support.
        return _defaults_only, {"Car Loan": 0.0}, 0.06, 0.0275, {}, {}
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)
    ws = wb["Main"]; wss = wb["Savings (Shared)"]
    def sf(v):
        try: return float(v) if v is not None else 0.0
        except (ValueError, TypeError): return 0.0
    net_pay_by_month = {}
    for col in range(3, 120):
        d = ws.cell(3, col).value
        if d is None: break
        net_pay_by_month[d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)] = sf(ws.cell(5, col).value)
    budget_rows = {"brokerage_main": 8, "btc": 10, "savings": 11, "income_brokerage": 9, "bonus": 15}
    excel_monthly = {}
    for col in range(3, 120):
        d = ws.cell(3, col).value
        if d is None: break
        key = d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)
        vals = {}
        for name, row in budget_rows.items():
            vals[name] = abs(sf(ws.cell(row, col).value))
        excel_monthly[key] = vals
    base_take_home = sf(ws.cell(5, 4).value) or 4430.0
    defaults = {
        "net_pay": base_take_home,
        "rent": abs(sf(ws.cell(6, 3).value)) or 1100.0,
        "car_insurance_gym": abs(sf(ws.cell(7, 3).value)) or 512.0,
        "food": abs(sf(ws.cell(12, 3).value)) or 700.0,
        "gas": abs(sf(ws.cell(13, 3).value)) or 150.0,
        "car_payment": abs(sf(ws.cell(14, 3).value)) or 1000.0,
        "brokerage_main": abs(sf(ws.cell(8, 3).value)) or 250.0,
        "btc": abs(sf(ws.cell(10, 3).value)) or 300.0,
        "savings": abs(sf(ws.cell(11, 3).value)) or 1000.0,
        "income_brokerage": abs(sf(ws.cell(9, 3).value)),
        "bonus": 0.0,
    }
    balances = {"Car Loan": sf(ws.cell(21, 2).value)}
    savings_yield = sf(wss.cell(1, 2).value) or 0.0275
    growth = sf(ws.cell(33, 2).value) or 0.06
    return defaults, balances, growth, savings_yield, net_pay_by_month, excel_monthly
# -- Build holdings -----------------------------------------------------------
def build_holdings_from_positions(eq_df, ot_df, cash):
    holdings_by_account = {}
    all_holdings = {}
    for df, section in [(eq_df, "equity"), (ot_df, "other")]:
        if df.empty or "Instrument" not in df.columns:
            continue
        for _, row in df.iterrows():
            ticker = str(row.get("Instrument", "")).strip()
            if not ticker: continue
            acct = str(row.get("Account Name", "")).strip()
            shares = row.get("Qty", 0)
            price = row.get("Mark", 0)
            value = row.get("Net Liq", shares * price)
            cost = row.get("Total Cost", 0)
            pl = row.get("P/L Open", 0)
            entry = {
                "ticker": ticker, "shares": shares, "price": price,
                "value": value, "cost": cost, "pl": pl, "account": acct,
                "yield_pct": KNOWN_YIELDS.get(ticker, 0.0), "section": section,
            }
            holdings_by_account.setdefault(acct, []).append(entry)
            all_holdings[f"{ticker}_{acct}"] = entry
    return holdings_by_account, all_holdings, cash
# -- Load data ----------------------------------------------------------------
defaults, liability_balances, growth_rate, savings_yield, net_pay_schedule, excel_monthly = load_excel_defaults()
eq_df_loaded, ot_df_loaded, cash_loaded, _raw_csv = load_latest_positions()
_hba_init, _ah_init, _cb_init = build_holdings_from_positions(eq_df_loaded, ot_df_loaded, cash_loaded)
# Keep module-level aliases for caches / functions defined before session state
holdings_by_account = _hba_init
all_holdings = _ah_init
cash_balance = _cb_init
# -- Live Price Fetching ------------------------------------------------------
@st.cache_data(ttl=300)
def fetch_live_prices(tickers):
    import yfinance as yf
    import concurrent.futures
    prices = {}
    fetchable = [t for t in tickers if t not in MONEY_MARKET_TICKERS and t not in MUTUAL_FUND_TICKERS]
    if not fetchable: return prices
    def _get(ticker):
        try:
            t = yf.Ticker(ticker)
            p = t.fast_info.get("lastPrice") or t.fast_info.get("previousClose")
            if p and not pd.isna(p): return ticker, float(p)
        except Exception: pass
        return ticker, None
    try:
        executor = concurrent.futures.ThreadPoolExecutor(max_workers=8)
        futures = {executor.submit(_get, t): t for t in fetchable}
        done, _ = concurrent.futures.wait(futures, timeout=10)
        for f in done:
            try:
                ticker, price = f.result(timeout=0)
                if price is not None: prices[ticker] = price
            except Exception: pass
        executor.shutdown(wait=False, cancel_futures=True)
    except Exception: pass
    return prices
def get_all_tickers():
    return {h["ticker"] for h in st.session_state.get("all_holdings", all_holdings).values()}
# -- Session State Init -------------------------------------------------------
# ── Retirement MUST be inited first so paycheck schedule uses live settings ──
if "retirement" not in st.session_state:
    _ret_defaults = {
        "roth_balance": 0.0, "k401_balance": 0.0,
        "current_age": 25, "retirement_age": 65,
        "monthly_roth_contribution": 0.0, "k401_contribution_pct": 5.0,
        "annual_return": 7.0, "hourly_rate": 39.50, "hours_per_year": 2080,
        "pay_frequency": "Biweekly", "benefits_per_check": 150.0, "state": "Florida",
        "next_payday": "2026-04-24",
    }
    if RETIREMENT_FILE.exists():
        try:
            _ret_defaults.update(json.loads(RETIREMENT_FILE.read_text()))
        except Exception:
            pass
    st.session_state.retirement = _ret_defaults
# Compute paycheck schedule from live session_state retirement settings
NET_PER_CHECK, PAYDAYS_PER_MONTH, paycheck_schedule = compute_paycheck_schedule(st.session_state.retirement)
PORTFOLIO_NAMES = sorted(holdings_by_account.keys())
def make_month_defaults(date_key):
    m = {
        "net_pay": defaults["net_pay"],
        "car_insurance_gym": defaults["car_insurance_gym"], "food": defaults["food"],
        "gas": defaults["gas"], "car_payment": defaults["car_payment"],
        "credit_card": 0.0,
        "brokerage_main": defaults["brokerage_main"], "btc": defaults["btc"],
        "savings": defaults["savings"], "income_brokerage": defaults["income_brokerage"],
        "roth_ira": 0.0,
        "bonus": 0.0, "other_income": 0.0, "one_time_expense": 0.0,
    }
    if date_key in paycheck_schedule:
        m["net_pay"] = paycheck_schedule[date_key]
    elif date_key in net_pay_schedule:
        m["net_pay"] = net_pay_schedule[date_key]
    if date_key in excel_monthly:
        for fld in INVEST_FIELDS + ["bonus"]:
            v = excel_monthly[date_key].get(fld, 0.0)
            if v: m[fld] = v
    return m
def build_div_holdings_from_positions():
    _ah = st.session_state.get("all_holdings", all_holdings)
    div_h = {}
    for key, h in _ah.items():
        if h["yield_pct"] > 0 and h["shares"] > 0:
            div_h[key] = {
                "ticker": h["ticker"], "shares": h["shares"], "price": h["price"],
                "yield_pct": h["yield_pct"], "account": h["account"],
            }
    return div_h
# ── Store holdings in session state so upload tab can refresh them ────────────
if "holdings_by_account" not in st.session_state:
    st.session_state.holdings_by_account = holdings_by_account
    st.session_state.all_holdings = all_holdings
    st.session_state.cash_balance = cash_balance
if "months" not in st.session_state:
    _budget_loaded = False
    if BUDGET_FILE.exists():
        try:
            _bdata = json.loads(BUDGET_FILE.read_text())
            st.session_state.months = _bdata["months"]
            if "car_paid_toggles" in _bdata:
                st.session_state.car_paid_toggles = _bdata["car_paid_toggles"]
            if "checks_received" in _bdata:
                st.session_state.checks_received = _bdata["checks_received"]
            _budget_loaded = True
        except Exception:
            pass
    if not _budget_loaded:
        now = datetime.now().replace(day=1)
        st.session_state.months = []
        for i in range(NUM_MONTHS):
            dt = now + relativedelta(months=i)
            date_key = dt.strftime("%Y-%m")
            st.session_state.months.append({
                "date": date_key, "label": dt.strftime("%b %Y"),
                **make_month_defaults(date_key),
            })
# Ensure all budget fields exist in every month (backward compat for new fields)
for _m in st.session_state.months:
    for _f in ALL_BUDGET_FIELDS:
        if _f not in _m:
            _m[_f] = 0.0
# ── Auto-roll: drop months in the past, append new ones at the end ──────────
_current_month_key = datetime.now().strftime("%Y-%m")
while st.session_state.months and st.session_state.months[0]["date"] < _current_month_key:
    st.session_state.months.pop(0)
while len(st.session_state.months) < NUM_MONTHS:
    _last = st.session_state.months[-1] if st.session_state.months else None
    if _last:
        _next_dt = datetime.strptime(_last["date"], "%Y-%m") + relativedelta(months=1)
    else:
        _next_dt = datetime.now().replace(day=1)
    _nk = _next_dt.strftime("%Y-%m")
    st.session_state.months.append({
        "date": _nk, "label": _next_dt.strftime("%b %Y"),
        **make_month_defaults(_nk),
    })
if "div_holdings" not in st.session_state:
    st.session_state.div_holdings = build_div_holdings_from_positions()
if "car_loan_balance" not in st.session_state:
    if CAR_LOAN_FILE.exists():
        st.session_state.car_loan_balance = json.loads(CAR_LOAN_FILE.read_text()).get(
            "balance", abs(liability_balances.get("Car Loan", 57000)))
    else:
        st.session_state.car_loan_balance = abs(liability_balances.get("Car Loan", 57000))
if "car_paid_toggles" not in st.session_state:
    st.session_state.car_paid_toggles = {}
if "checks_received" not in st.session_state:
    st.session_state.checks_received = {}

if "credit_card_balance" not in st.session_state:
    if CREDIT_CARD_BAL_FILE.exists():
        st.session_state.credit_card_balance = json.loads(CREDIT_CARD_BAL_FILE.read_text()).get("balance", 0.0)
    else:
        st.session_state.credit_card_balance = 0.0
if "portfolio_contributions" not in st.session_state:
    st.session_state.portfolio_contributions = {acct: [0.0] * NUM_MONTHS for acct in PORTFOLIO_NAMES}
if "growth_rate" not in st.session_state:
    st.session_state.growth_rate = float(growth_rate)
if "savings_yield" not in st.session_state:
    st.session_state.savings_yield = float(savings_yield)
if "portfolio_allocations" not in st.session_state:
    non_shared = [a for a in PORTFOLIO_NAMES if a != "Shared"]
    st.session_state.portfolio_allocations = {acct: 0.0 for acct in PORTFOLIO_NAMES}
    for a in non_shared:
        st.session_state.portfolio_allocations[a] = round(100.0 / max(len(non_shared), 1), 1)
# retirement was already initialized above (before paycheck schedule)
if "share_alloc_pcts" not in st.session_state:
    st.session_state.share_alloc_pcts = {}
    for acct, holdings in st.session_state.holdings_by_account.items():
        non_mm = [h["ticker"] for h in holdings if h["ticker"] not in MONEY_MARKET_TICKERS]
        if non_mm:
            st.session_state.share_alloc_pcts[acct] = {non_mm[0]: 100.0}
        elif holdings:
            st.session_state.share_alloc_pcts[acct] = {holdings[0]["ticker"]: 100.0}
        else:
            st.session_state.share_alloc_pcts[acct] = {}
    if PROJECTION_FILE.exists():
        saved = json.loads(PROJECTION_FILE.read_text())
        if "share_alloc_pcts" in saved:
            st.session_state.share_alloc_pcts.update(saved["share_alloc_pcts"])
        elif "share_alloc_tickers" in saved:
            for a, t in saved["share_alloc_tickers"].items():
                st.session_state.share_alloc_pcts[a] = {t: 100.0}
# -- CSS ----------------------------------------------------------------------
st.markdown("""
<style>
    :root {
        --app-blue: #dceeff;
        --app-blue-strong: #8ec5ff;
        --app-blue-deep: #2d6fb5;
        --app-panel: #ffffff;
        --app-text: #111827;
        --app-muted: #58728d;
        --app-border: #c2def6;
    }
    /* Hide Streamlit header black bar */
    header[data-testid="stHeader"] {
        background: transparent !important;
        height: 0 !important;
        min-height: 0 !important;
    }
    #MainMenu, footer { visibility: hidden; }
    .stApp {
        background: linear-gradient(180deg, #f7fbff 0%, #edf6ff 100%);
        color: var(--app-text);
    }
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #ffffff 0%, #eef6ff 100%);
        border-right: 1px solid var(--app-border);
    }
    section[data-testid="stSidebar"] * { color: var(--app-text) !important; }
    /* Metric cards */
    div[data-testid="stMetric"] {
        background: rgba(255, 255, 255, 0.92);
        border: 1px solid var(--app-border);
        border-radius: 16px;
        padding: 10px 14px;
        height: 110px;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        box-shadow: 0 8px 30px rgba(76, 129, 189, 0.08);
        overflow: hidden;
    }
    div[data-testid="stMetric"] label,
    div[data-testid="stMetric"] [data-testid="stMetricLabel"] p,
    div[data-testid="stMetric"] [data-testid="stMetricValue"],
    div[data-testid="stMetric"] [data-testid="stMetricDelta"] {
        color: var(--app-text) !important;
    }
    /* DataFrames */
    div[data-testid="stDataFrame"] { background: var(--app-panel); }
    div[data-testid="stDataFrame"] * { color: var(--app-text) !important; }
    /* Inputs */
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stTextArea"] textarea,
    div[data-testid="stDateInput"] input {
        color: var(--app-text) !important;
        background-color: #f7fbff !important;
        border: 2px solid #8fb7dd !important;
        border-radius: 8px !important;
    }
    div[data-testid="stTextInput"],
    div[data-testid="stNumberInput"],
    div[data-testid="stTextArea"],
    div[data-testid="stDateInput"],
    div[data-testid="stSelectbox"] {
        border: 2px solid #9ec2e4;
        border-radius: 12px;
        background: #ffffff;
        padding: 2px 8px 4px 8px;
        box-shadow: 0 1px 0 rgba(45, 111, 181, 0.08);
    }
    div[data-testid="stSelectbox"] > div > div,
    div[data-testid="stSelectbox"] span {
        color: var(--app-text) !important;
        background-color: #f7fbff !important;
        border: none !important;
    }
    /* HIDE all +/- stepper buttons */
    div[data-testid="stNumberInput"] button { display: none !important; }
    /* Labels */
    label, .stMarkdown p, p { color: var(--app-text) !important; }
    /* Tabs - styled as bubbles */
    div[data-baseweb="tab-list"] {
        gap: 6px !important;
        background: transparent !important;
        border-bottom: none !important;
        display: flex !important;
        flex-wrap: wrap !important;
    }
    button[data-baseweb="tab"] {
        background: #f0f7ff !important;
        border: 1.5px solid #c2def6 !important;
        border-radius: 12px !important;
        color: var(--app-text) !important;
        padding: 8px 16px !important;
        font-weight: 600 !important;
        box-shadow: none !important;
        min-height: 36px !important;
        display: inline-flex !important;
        align-items: center !important;
        visibility: visible !important;
        opacity: 1 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        background: #2d6fb5 !important;
        border-color: #2d6fb5 !important;
        color: #ffffff !important;
    }
    /* CHECKBOXES: white box, blue check */
    [data-baseweb="checkbox"] > div {
        background-color: #ffffff !important;
        border-color: var(--app-blue-deep) !important;
        border-radius: 4px !important;
    }
    [data-baseweb="checkbox"] input:checked + div,
    [data-baseweb="checkbox"] input:checked ~ div {
        background-color: var(--app-blue-deep) !important;
        border-color: var(--app-blue-deep) !important;
    }
    [data-testid="stCheckbox"] label span,
    [data-testid="stCheckbox"] label p { color: var(--app-text) !important; }
    /* RADIO BUTTONS: white bubble, blue fill */
    [data-baseweb="radio"] > div {
        background-color: #ffffff !important;
        border-color: var(--app-blue-deep) !important;
    }
    [data-baseweb="radio"] input:checked + div,
    [data-baseweb="radio"] input:checked ~ div { border-color: var(--app-blue-deep) !important; }
    [data-baseweb="radio"] input:checked + div div,
    [data-baseweb="radio"] input:checked ~ div div { background-color: var(--app-blue-deep) !important; }
    [data-testid="stRadio"] label span { color: var(--app-text) !important; }
    /* Sidebar navigation: bubble-style selector */
    /* Hide just the radio circle dot, keep text visible */
    /* Sidebar navigation: highlight selected item, hide nothing */
    section[data-testid="stSidebar"] div[role="radiogroup"] label {
        border-radius: 10px !important;
        padding: 5px 10px !important;
        margin: 2px 0 !important;
        cursor: pointer !important;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label:has([role="radio"][aria-checked="true"]) {
        background: #2d6fb5 !important;
        border-radius: 10px !important;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label:has([role="radio"][aria-checked="true"]) p,
    section[data-testid="stSidebar"] div[role="radiogroup"] label:has([role="radio"][aria-checked="true"]) span {
        color: #ffffff !important;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] [role="radio"][aria-checked="true"] > div {
        background: #ffffff !important;
        border-color: #ffffff !important;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] *:focus,
    section[data-testid="stSidebar"] div[role="radiogroup"] *:focus-visible {
        outline: none !important;
        box-shadow: none !important;
    }
    /* Share allocation summary blocks */
    .alloc-total-box {
        border: 2px solid #9ec2e4;
        border-radius: 12px;
        background: #ffffff;
        padding: 10px 12px;
        min-height: 88px;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    .alloc-total-label {
        font-size: 0.9rem;
        font-weight: 700;
        color: #111827;
    }
    .alloc-total-value {
        font-size: 1.75rem;
        font-weight: 800;
        line-height: 1.1;
        margin-top: 4px;
    }
    .alloc-total-rem {
        font-size: 0.9rem;
        color: #58728d;
        margin-top: 2px;
    }
    .money-pos { font-size: 2.8rem; font-weight: 700; color: #1f8f58 !important; text-align: center; }
    .money-neg { font-size: 2.8rem; font-weight: 700; color: #d3465a !important; text-align: center; }
    .money-label { font-size: 1rem; color: var(--app-muted) !important; text-align: center; margin-bottom: 0; }
    .section-hdr { font-size: 1.1rem; font-weight: 700; border-bottom: 2px solid var(--app-blue-strong);
                   padding-bottom: 4px; margin-top: 1rem; color: var(--app-blue-deep) !important; }
    .summary-card {
        background: rgba(255, 255, 255, 0.95);
        border: 1px solid var(--app-border);
        border-radius: 18px;
        padding: 18px;
        box-shadow: 0 10px 28px rgba(72, 126, 184, 0.08);
        color: var(--app-text);
    }
    .summary-card * { color: var(--app-text); }
    .summary-accent {
        color: var(--app-blue-deep) !important;
        font-size: 0.9rem; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.08em;
    }
    div[data-testid="stNumberInput"] { margin-bottom: -10px; }
    .stButton > button, div[data-testid="stFormSubmitButton"] > button {
        background: linear-gradient(180deg, #edf6ff 0%, #d5ebff 100%);
        color: #103b63 !important;
        border: 1px solid var(--app-blue-strong);
        border-radius: 12px; font-weight: 700;
    }
    div[data-testid="stCaptionContainer"] p { color: var(--app-muted) !important; }
    details summary span { color: var(--app-text) !important; }
</style>
""", unsafe_allow_html=True)


def render_overview(include_shared, live_prices):
    st.title("Financial Overview")
    months = st.session_state.months
    current_month = months[0] if months else {"label": datetime.now().strftime("%b %Y")}
    current_totals = compute_month_totals(current_month) if months else {
        "income_total": 0.0, "expense_total": 0.0, "invested_total": 0.0, "money_left": 0.0,
    }
    _cb = st.session_state.get("cash_balance", 0.0)
    _ah = st.session_state.get("all_holdings", {})
    net_worth = -abs(st.session_state.car_loan_balance) + _cb
    for holding in _ah.values():
        if holding["account"] == "Shared" and not include_shared:
            continue
        ticker = holding["ticker"]
        net_worth += holding["value"] if ticker in MONEY_MARKET_TICKERS else holding["shares"] * live_prices.get(ticker, holding["price"])
    hero1, hero2, hero3 = st.columns(3)
    hero1.metric("Current Net Worth", f"${net_worth:,.0f}")
    hero2.metric(f"{current_month['label']} Money Left", f"${current_totals['money_left']:,.0f}")
    hero3.metric(f"{current_month['label']} Invested", f"${current_totals['invested_total']:,.0f}")
    st.markdown("---")
    c1, c2 = st.columns([1.25, 1.0])
    with c1:
        st.markdown(
            f"""
            <div class="summary-card">
                <div class="summary-accent">Current Month</div>
                <h2 style="margin:0.35rem 0 0.8rem 0; color:#111827;">{current_month['label']} Budget Snapshot</h2>
                <div style="display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px;">
                    <div><div class="summary-accent">Income</div><div style="font-size:1.8rem;font-weight:700;">${current_totals['income_total']:,.0f}</div></div>
                    <div><div class="summary-accent">Expenses</div><div style="font-size:1.8rem;font-weight:700;">${current_totals['expense_total']:,.0f}</div></div>
                    <div><div class="summary-accent">Investments</div><div style="font-size:1.8rem;font-weight:700;">${current_totals['invested_total']:,.0f}</div></div>
                    <div><div class="summary-accent">Left Over</div><div style="font-size:1.8rem;font-weight:700;color:{'#1f8f58' if current_totals['money_left'] >= 0 else '#d3465a'};">${current_totals['money_left']:,.0f}</div></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f"""
            <div class="summary-card">
                <div class="summary-accent">Current Month Line Items</div>
                <div style="margin-top:0.85rem;display:grid;gap:0.55rem;">
                    <div><strong>Net Pay:</strong> ${float(current_month.get('net_pay', 0.0)):,.0f}</div>
                    <div><strong>Food:</strong> ${float(current_month.get('food', 0.0)):,.0f}</div>
                    <div><strong>Gas:</strong> ${float(current_month.get('gas', 0.0)):,.0f}</div>
                    <div><strong>Car Payment:</strong> ${float(current_month.get('car_payment', 0.0)):,.0f}</div>
                    <div><strong>Brokerage:</strong> ${float(current_month.get('brokerage_main', 0.0)):,.0f}</div>
                    <div><strong>BTC:</strong> ${float(current_month.get('btc', 0.0)):,.0f}</div>
                    <div><strong>Savings:</strong> ${float(current_month.get('savings', 0.0)):,.0f}</div>
                    <div><strong>Income Brokerage:</strong> ${float(current_month.get('income_brokerage', 0.0)):,.0f}</div>
                    <div><strong>Roth IRA:</strong> ${float(current_month.get('roth_ira', 0.0)):,.0f}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
# -- Last Saved Timestamp -----------------------------------------------------
_last_ts = ""
if LAST_SAVED_FILE.exists():
    try:
        _last_ts = json.loads(LAST_SAVED_FILE.read_text()).get("ts", "")
    except Exception:
        pass
if _last_ts:
    st.sidebar.caption(f"Last saved: {_last_ts}")
else:
    st.sidebar.caption("Not yet saved")
# -- Sidebar ------------------------------------------------------------------
page = st.sidebar.radio(
    "Navigate",
    ["Overview", "12-Month Budget", "Portfolio Projection", "Analytics",
     "Dividend Forecaster", "Retirement & Paycheck", "Balance Sheet",
     "Upload Positions"],
    label_visibility="collapsed",
)
st.sidebar.markdown("---")
include_shared = st.sidebar.toggle("Include Shared Portfolio", value=False, key="incl_shared")
st.sidebar.markdown("---")
st.sidebar.markdown("**Car Loan Balance**")
new_car_bal = st.sidebar.number_input(
    "Car Loan Owed", value=float(st.session_state.car_loan_balance),
    step=100.0, key="car_loan_input", format="%.0f",
)
if new_car_bal != st.session_state.car_loan_balance:
    st.session_state.car_loan_balance = new_car_bal
    CAR_LOAN_FILE.write_text(json.dumps({"balance": new_car_bal}))
st.sidebar.markdown("**Growth Rate**")
growth_input = st.sidebar.number_input(
    "Annual Growth %", value=float(st.session_state.growth_rate * 100.0),
    step=0.25, min_value=-50.0, max_value=100.0, key="growth_rate_input", format="%.2f",
)
st.session_state.growth_rate = growth_input / 100.0
st.sidebar.caption(f"Savings Yield: {st.session_state.savings_yield:.2%}")
st.sidebar.markdown("---")
use_live_prices = st.sidebar.toggle("Use Live Market Prices", value=False, key="use_live_prices")
sc1, sc2 = st.sidebar.columns(2)
if sc1.button("\U0001f4be Save All"):
    save_all()
    st.sidebar.success("All saved!")
if sc2.button("\U0001f4c2 Load All"):
    load_all()
    st.sidebar.success("All loaded!")
if st.sidebar.button("Apply Changes", type="primary", use_container_width=True):
    save_all()
    st.cache_data.clear()
    st.rerun()
# -- Live prices + Net Worth --------------------------------------------------
st.sidebar.markdown("---")
all_tickers = get_all_tickers()
fetchable_tickers = tuple(sorted(
    t for t in all_tickers if t not in MONEY_MARKET_TICKERS and t not in MUTUAL_FUND_TICKERS
))
_pages_needing_live = {"Overview", "Portfolio Projection", "Analytics", "Dividend Forecaster", "Retirement & Paycheck", "Balance Sheet"}
live_prices = fetch_live_prices(fetchable_tickers) if use_live_prices and fetchable_tickers and page in _pages_needing_live else {}
def live_net_worth():
    _cb = st.session_state.get("cash_balance", cash_balance)
    _ah = st.session_state.get("all_holdings", {})
    nw = -abs(st.session_state.car_loan_balance) + _cb
    for key, h in _ah.items():
        if not include_shared and h["account"] == "Shared":
            continue
        t = h["ticker"]
        nw += h["value"] if t in MONEY_MARKET_TICKERS else h["shares"] * live_prices.get(t, h["price"])
    return nw
live_nw = live_net_worth()
st.sidebar.metric("Current Net Worth", f"${live_nw:,.0f}")
if live_prices:
    st.sidebar.caption(f"Live prices for {len(live_prices)}/{len(fetchable_tickers)} tickers")
else:
    st.sidebar.caption("Using uploaded position statement prices")
st.sidebar.caption(f"Growth: {st.session_state.growth_rate:.2%} | Sav Yield: {st.session_state.savings_yield:.2%}")
# -- Page routing -------------------------------------------------------------
if page == "Overview":
    render_overview(include_shared, live_prices)
elif page == "12-Month Budget":
    from tabs.budget import render
    _npc, _pdm, _ = compute_paycheck_schedule(st.session_state.retirement)
    render(include_shared, net_per_check=_npc, paydays_per_month=_pdm)
elif page == "Portfolio Projection":
    from tabs.projection import render
    _hba = st.session_state.get("holdings_by_account", {})
    _ah = st.session_state.get("all_holdings", {})
    _cb = st.session_state.get("cash_balance", 0.0)
    render(_hba, _ah, live_prices, _cb, st.session_state.growth_rate, st.session_state.savings_yield, include_shared, sorted(_hba.keys()))
elif page == "Analytics":
    from tabs.analytics import render
    _hba = st.session_state.get("holdings_by_account", {})
    _ah = st.session_state.get("all_holdings", {})
    _cb = st.session_state.get("cash_balance", 0.0)
    render(_hba, _ah, live_prices, _cb, st.session_state.growth_rate, st.session_state.savings_yield, include_shared, sorted(_hba.keys()))
elif page == "Dividend Forecaster":
    from tabs.dividends import render
    render(include_shared, live_prices, st.session_state.growth_rate)
elif page == "Retirement & Paycheck":
    from tabs.retirement import render
    _hba = st.session_state.get("holdings_by_account", {})
    _cb = st.session_state.get("cash_balance", 0.0)
    render(_hba, live_prices, _cb, st.session_state.growth_rate, st.session_state.savings_yield, include_shared)
elif page == "Balance Sheet":
    from tabs.balance_sheet import render
    _hba = st.session_state.get("holdings_by_account", {})
    _ah = st.session_state.get("all_holdings", {})
    _cb = st.session_state.get("cash_balance", 0.0)
    render(_hba, _ah, live_prices, _cb, include_shared)
elif page == "Upload Positions":
    from tabs.upload import render
    render(parse_position_csv, build_holdings_from_positions, load_latest_positions,
           eq_df_loaded, ot_df_loaded, POSITION_DIR)