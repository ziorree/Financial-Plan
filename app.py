"""Financial Planner - main entry point.
Loads data, renders sidebar, routes to per-tab modules.
"""
import streamlit as st
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import csv
import json
# -- Config -------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "Main Financial Plan.xlsx"
POSITION_DIR = BASE_DIR / "Position Files"
st.set_page_config(page_title="Financial Planner", layout="wide", page_icon="\U0001f4b0")
from tabs.shared import (
    MONEY_MARKET_TICKERS, MUTUAL_FUND_TICKERS, KNOWN_YIELDS,
    EXPENSE_FIELDS, INVEST_FIELDS, INCOME_FIELDS, NUM_MONTHS,
    ROW_LABELS, INVEST_TO_PORTFOLIO, ALL_BUDGET_FIELDS,
    CAR_LOAN_FILE, RETIREMENT_FILE, PROJECTION_FILE, CREDIT_CARD_BAL_FILE, LAST_SAVED_FILE,
    BUDGET_FILE,
    save_all, load_all, calc_annual_taxes,
)
# -- Parse Position CSV -------------------------------------------------------
def parse_dollar(s):
    if s is None:
        return 0.0
    s = str(s).replace("$", "").replace(",", "").replace("(", "-").replace(")", "").strip()
    if s in ("", "N/A", "<empty>"):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0
def parse_position_csv(filepath):
    content = Path(filepath).read_text(encoding="utf-8-sig")
    lines = content.splitlines()
    equities, others = [], []
    section = None
    header = None
    cash_amount = 0.0
    for line in lines:
        row = next(csv.reader([line]))
        if not row or not row[0].strip():
            continue
        first = row[0].strip()
        if first.startswith("Position Statement"):
            continue
        if first == "Equities and Equity Options":
            section = "equities"; continue
        if first == "Others":
            section = "others"; continue
        if first == "Instrument" and section:
            header = [c.strip() for c in row]; continue
        if first.startswith("Cash & Sweep"):
            for part in row:
                v = parse_dollar(part)
                if v != 0:
                    cash_amount = v
            section = None; continue
        skips = ("OVERALL", "BP ", "OVERNIGHT", "Subtotal", "Overall", "P/L", "Available")
        if any(first.startswith(s) for s in skips):
            continue
        if header and section:
            if len(first) <= 6 and first.replace(".", "").isalpha():
                record = {}
                for i, h in enumerate(header):
                    record[h] = row[i].strip() if i < len(row) else ""
                (equities if section == "equities" else others).append(record)
    def build_df(records):
        if not records:
            return pd.DataFrame()
        df = pd.DataFrame(records)
        if "Qty" in df.columns:
            df["Qty"] = df["Qty"].apply(lambda x: parse_dollar(str(x).replace("+", "")))
        for col in ["Trade Price", "Mark", "Net Liq", "Total Cost", "P/L Open", "P/L Day", "Mrk Chng"]:
            if col in df.columns:
                df[col] = df[col].apply(parse_dollar)
        return df
    return build_df(equities), build_df(others), cash_amount, content
@st.cache_data
def load_latest_positions():
    if not POSITION_DIR.exists():
        return pd.DataFrame(), pd.DataFrame(), 0.0, ""
    csvs = sorted(POSITION_DIR.glob("*.csv"))
    if not csvs:
        return pd.DataFrame(), pd.DataFrame(), 0.0, ""
    return parse_position_csv(csvs[-1])
# -- Load Excel defaults ------------------------------------------------------
@st.cache_data
def load_excel_defaults():
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)
    ws = wb["Main"]; wss = wb["Savings (Shared)"]
    def sf(v):
        try: return float(v) if v is not None else 0.0
        except (ValueError, TypeError): return 0.0
    net_pay_by_month = {}
    for col in range(3, 120):
        d = ws.cell(3, col).value
        if d is None: break
        net_pay_by_month[d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)] = sf(ws.cell(5, col).value)
    budget_rows = {"brokerage_main": 8, "btc": 10, "savings": 11, "income_brokerage": 9, "bonus": 15}
    excel_monthly = {}
    for col in range(3, 120):
        d = ws.cell(3, col).value
        if d is None: break
        key = d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)
        vals = {}
        for name, row in budget_rows.items():
            vals[name] = abs(sf(ws.cell(row, col).value))
        excel_monthly[key] = vals
    base_take_home = sf(ws.cell(5, 4).value) or 4430.0
    defaults = {
        "net_pay": base_take_home,
        "rent": abs(sf(ws.cell(6, 3).value)) or 1100.0,
        "car_insurance_gym": abs(sf(ws.cell(7, 3).value)) or 512.0,
        "food": abs(sf(ws.cell(12, 3).value)) or 700.0,
        "gas": abs(sf(ws.cell(13, 3).value)) or 150.0,
        "car_payment": abs(sf(ws.cell(14, 3).value)) or 1000.0,
        "brokerage_main": abs(sf(ws.cell(8, 3).value)) or 250.0,
        "btc": abs(sf(ws.cell(10, 3).value)) or 300.0,
        "savings": abs(sf(ws.cell(11, 3).value)) or 1000.0,
        "income_brokerage": abs(sf(ws.cell(9, 3).value)),
        "bonus": 0.0,
    }
    balances = {"Car Loan": sf(ws.cell(21, 2).value)}
    savings_yield = sf(wss.cell(1, 2).value) or 0.0275
    growth = sf(ws.cell(33, 2).value) or 0.06
    return defaults, balances, growth, savings_yield, net_pay_by_month, excel_monthly
# -- Build holdings -----------------------------------------------------------
def build_holdings_from_positions(eq_df, ot_df, cash):
    holdings_by_account = {}
    all_holdings = {}
    for df, section in [(eq_df, "equity"), (ot_df, "other")]:
        if df.empty or "Instrument" not in df.columns:
            continue
        for _, row in df.iterrows():
            ticker = str(row.get("Instrument", "")).strip()
            if not ticker: continue
            acct = str(row.get("Account Name", "")).strip()
            shares = row.get("Qty", 0)
            price = row.get("Mark", 0)
            value = row.get("Net Liq", shares * price)
            cost = row.get("Total Cost", 0)
            pl = row.get("P/L Open", 0)
            entry = {
                "ticker": ticker, "shares": shares, "price": price,
                "value": value, "cost": cost, "pl": pl, "account": acct,
                "yield_pct": KNOWN_YIELDS.get(ticker, 0.0), "section": section,
            }
            holdings_by_account.setdefault(acct, []).append(entry)
            all_holdings[f"{ticker}_{acct}"] = entry
    return holdings_by_account, all_holdings, cash
# -- Load data ----------------------------------------------------------------
try:
    defaults, liability_balances, growth_rate, savings_yield, net_pay_schedule, excel_monthly = load_excel_defaults()
except Exception:
    defaults = {"net_pay": 4430.0, "car_insurance_gym": 512.0, "food": 700.0, "gas": 150.0, "car_payment": 1000.0, "brokerage_main": 250.0, "btc": 300.0, "savings": 1000.0, "income_brokerage": 0.0, "bonus": 0.0}
    liability_balances = {"Car Loan": 57000}
    growth_rate = 0.06
    savings_yield = 0.0275
    net_pay_schedule = {}
    excel_monthly = {}

eq_df_loaded, ot_df_loaded, cash_loaded, _raw_csv = load_latest_positions()
holdings_by_account, all_holdings, cash_balance = build_holdings_from_positions(eq_df_loaded, ot_df_loaded, cash_loaded)
# -- Live Price Fetching ------------------------------------------------------
@st.cache_data(ttl=300)
def fetch_live_prices(tickers):
    import yfinance as yf
    import concurrent.futures
    prices = {}
    fetchable = [t for t in tickers if t not in MONEY_MARKET_TICKERS and t not in MUTUAL_FUND_TICKERS]
    if not fetchable: return prices
    def _get(ticker):
        try:
            t = yf.Ticker(ticker)
            p = t.fast_info.get("lastPrice") or t.fast_info.get("previousClose")
            if p and not pd.isna(p): return ticker, float(p)
        except Exception: pass
        return ticker, None
    try:
        executor = concurrent.futures.ThreadPoolExecutor(max_workers=8)
        futures = {executor.submit(_get, t): t for t in fetchable}
        done, _ = concurrent.futures.wait(futures, timeout=10)
        for f in done:
            try:
                ticker, price = f.result(timeout=0)
                if price is not None: prices[ticker] = price
            except Exception: pass
        executor.shutdown(wait=False, cancel_futures=True)
    except Exception: pass
    return prices
def get_all_tickers():
    return {h["ticker"] for h in all_holdings.values()}
# -- Session State Init -------------------------------------------------------
PORTFOLIO_NAMES = sorted(holdings_by_account.keys())
# Compute paycheck-based net pay per month from saved retirement settings
def _compute_paycheck_schedule():
    """Returns (net_per_check, {date_key: [payday_date, ...]}, {date_key: total_net})"""
    ret = {}
    if RETIREMENT_FILE.exists():
        try:
            ret = json.loads(RETIREMENT_FILE.read_text())
        except Exception:
            return 0.0, {}, {}
    if not ret.get("hourly_rate"):
        return 0.0, {}, {}
    gross_annual = ret["hourly_rate"] * ret.get("hours_per_year", 2080)
    freq = ret.get("pay_frequency", "Biweekly")
    checks = {"Biweekly": 26, "Semi-Monthly": 24, "Monthly": 12}.get(freq, 26)
    gross_pc = gross_annual / checks
    k401_pct = ret.get("k401_contribution_pct", 5.0) / 100.0
    k401_pc = gross_pc * k401_pct
    k401_annual = gross_annual * k401_pct
    benefits = ret.get("benefits_per_check", 0)
    fed, state_tax, fica = calc_annual_taxes(gross_annual, ret.get("state", "Florida"), k401_annual)
    net_pc = gross_pc - k401_pc - benefits - fed / checks - state_tax / checks - fica / checks
    next_pd = ret.get("next_payday", "2026-04-24")
    try:
        start = datetime.strptime(next_pd, "%Y-%m-%d").date()
    except Exception:
        return 0.0, {}, {}
    delta = timedelta(days=14) if freq == "Biweekly" else timedelta(days=15)
    paydays_per_month = {}
    current = start
    end = start + timedelta(days=400)
    while current <= end:
        key = current.strftime("%Y-%m")
        paydays_per_month.setdefault(key, []).append(current)
        current += delta
    totals = {k: round(len(v) * net_pc, 2) for k, v in paydays_per_month.items()}
    return round(net_pc, 2), paydays_per_month, totals
NET_PER_CHECK, PAYDAYS_PER_MONTH, paycheck_schedule = _compute_paycheck_schedule()
def make_month_defaults(date_key):
    m = {
        "net_pay": defaults["net_pay"],
        "car_insurance_gym": defaults["car_insurance_gym"], "food": defaults["food"],
        "gas": defaults["gas"], "car_payment": defaults["car_payment"],
        "credit_card": 0.0,
        "brokerage_main": defaults["brokerage_main"], "btc": defaults["btc"],
        "savings": defaults["savings"], "income_brokerage": defaults["income_brokerage"],
        "roth_ira": 0.0,
        "bonus": 0.0, "other_income": 0.0, "one_time_expense": 0.0,
    }
    if date_key in paycheck_schedule:
        m["net_pay"] = paycheck_schedule[date_key]
    elif date_key in net_pay_schedule:
        m["net_pay"] = net_pay_schedule[date_key]
    if date_key in excel_monthly:
        for fld in INVEST_FIELDS + ["bonus"]:
            v = excel_monthly[date_key].get(fld, 0.0)
            if v: m[fld] = v
    return m
def build_div_holdings_from_positions():
    div_h = {}
    for key, h in all_holdings.items():
        if h["yield_pct"] > 0 and h["shares"] > 0:
            div_h[key] = {
                "ticker": h["ticker"], "shares": h["shares"], "price": h["price"],
                "yield_pct": h["yield_pct"], "account": h["account"],
            }
    return div_h
if "months" not in st.session_state:
    now = datetime.now().replace(day=1)
    st.session_state.months = []
    for i in range(NUM_MONTHS):
        dt = now + relativedelta(months=i)
        date_key = dt.strftime("%Y-%m")
        st.session_state.months.append({
            "date": date_key, "label": dt.strftime("%b %Y"),
            **make_month_defaults(date_key),
        })
# Ensure all budget fields exist in every month (backward compat for new fields)
for _m in st.session_state.months:
    for _f in ALL_BUDGET_FIELDS:
        if _f not in _m:
            _m[_f] = 0.0
# ── Auto-roll: drop months in the past, append new ones at the end ──────────
_current_month_key = datetime.now().strftime("%Y-%m")
while st.session_state.months and st.session_state.months[0]["date"] < _current_month_key:
    st.session_state.months.pop(0)
while len(st.session_state.months) < NUM_MONTHS:
    _last = st.session_state.months[-1] if st.session_state.months else None
    if _last:
        _next_dt = datetime.strptime(_last["date"], "%Y-%m") + relativedelta(months=1)
    else:
        _next_dt = datetime.now().replace(day=1)
    _nk = _next_dt.strftime("%Y-%m")
    st.session_state.months.append({
        "date": _nk, "label": _next_dt.strftime("%b %Y"),
        **make_month_defaults(_nk),
    })
if "div_holdings" not in st.session_state:
    st.session_state.div_holdings = build_div_holdings_from_positions()
if "car_loan_balance" not in st.session_state:
    if CAR_LOAN_FILE.exists():
        st.session_state.car_loan_balance = json.loads(CAR_LOAN_FILE.read_text()).get(
            "balance", abs(liability_balances.get("Car Loan", 57000)))
    else:
        st.session_state.car_loan_balance = abs(liability_balances.get("Car Loan", 57000))
if "car_paid_toggles" not in st.session_state:
    st.session_state.car_paid_toggles = {}
if "checks_received" not in st.session_state:
    st.session_state.checks_received = {}
if "credit_card_balance" not in st.session_state:
    if CREDIT_CARD_BAL_FILE.exists():
        st.session_state.credit_card_balance = json.loads(CREDIT_CARD_BAL_FILE.read_text()).get("balance", 0.0)
    else:
        st.session_state.credit_card_balance = 0.0
if "portfolio_contributions" not in st.session_state:
    st.session_state.portfolio_contributions = {acct: [0.0] * NUM_MONTHS for acct in PORTFOLIO_NAMES}
if "portfolio_allocations" not in st.session_state:
    non_shared = [a for a in PORTFOLIO_NAMES if a != "Shared"]
    st.session_state.portfolio_allocations = {acct: 0.0 for acct in PORTFOLIO_NAMES}
    for a in non_shared:
        st.session_state.portfolio_allocations[a] = round(100.0 / max(len(non_shared), 1), 1)
if "retirement" not in st.session_state:
    ret_defaults = {
        "roth_balance": 0.0, "k401_balance": 0.0,
        "current_age": 25, "retirement_age": 65,
        "monthly_roth_contribution": 0.0, "k401_contribution_pct": 5.0,
        "annual_return": 7.0, "hourly_rate": 39.50, "hours_per_year": 2080,
        "pay_frequency": "Biweekly", "benefits_per_check": 150.0, "state": "Florida",
        "next_payday": "2026-04-24",
    }
    if RETIREMENT_FILE.exists():
        ret_defaults.update(json.loads(RETIREMENT_FILE.read_text()))
    st.session_state.retirement = ret_defaults
if "share_alloc_pcts" not in st.session_state:
    st.session_state.share_alloc_pcts = {}
    for acct, holdings in holdings_by_account.items():
        non_mm = [h["ticker"] for h in holdings if h["ticker"] not in MONEY_MARKET_TICKERS]
        if non_mm:
            st.session_state.share_alloc_pcts[acct] = {non_mm[0]: 100.0}
        elif holdings:
            st.session_state.share_alloc_pcts[acct] = {holdings[0]["ticker"]: 100.0}
        else:
            st.session_state.share_alloc_pcts[acct] = {}
    if PROJECTION_FILE.exists():
        saved = json.loads(PROJECTION_FILE.read_text())
        if "share_alloc_pcts" in saved:
            st.session_state.share_alloc_pcts.update(saved["share_alloc_pcts"])
        elif "share_alloc_tickers" in saved:
            for a, t in saved["share_alloc_tickers"].items():
                st.session_state.share_alloc_pcts[a] = {t: 100.0}
# -- CSS ----------------------------------------------------------------------
st.markdown("""
<style>
    .money-pos { font-size: 2.8rem; font-weight: 700; color: #2ecc71; text-align: center; }
    .money-neg { font-size: 2.8rem; font-weight: 700; color: #e74c3c; text-align: center; }
    .money-label { font-size: 1rem; color: #888; text-align: center; margin-bottom: 0; }
    .section-hdr { font-size: 1.1rem; font-weight: 600; border-bottom: 2px solid #444;
                   padding-bottom: 4px; margin-top: 1rem; }
    div[data-testid="stNumberInput"] { margin-bottom: -10px; }
</style>
""", unsafe_allow_html=True)
# -- Last Saved Timestamp -----------------------------------------------------
_last_ts = ""
if LAST_SAVED_FILE.exists():
    try:
        _last_ts = json.loads(LAST_SAVED_FILE.read_text()).get("ts", "")
    except Exception:
        pass
if _last_ts:
    st.sidebar.caption(f"Last saved: {_last_ts}")
else:
    st.sidebar.caption("Not yet saved")
# -- Sidebar ------------------------------------------------------------------
page = st.sidebar.radio(
    "Navigate",
    ["12-Month Budget", "Portfolio Projection", "Analytics",
     "Dividend Forecaster", "Retirement & Paycheck", "Balance Sheet",
     "Upload Positions"],
    label_visibility="collapsed",
)
st.sidebar.markdown("---")
include_shared = st.sidebar.toggle("Include Shared Portfolio", value=False, key="incl_shared")
st.sidebar.markdown("---")
st.sidebar.markdown("**Car Loan Balance**")
new_car_bal = st.sidebar.number_input(
    "Car Loan Owed", value=float(st.session_state.car_loan_balance),
    step=100.0, key="car_loan_input", format="%.0f",
)
if new_car_bal != st.session_state.car_loan_balance:
    st.session_state.car_loan_balance = new_car_bal
    from tabs.shared import CAR_LOAN_FILE
    CAR_LOAN_FILE.write_text(json.dumps({"balance": new_car_bal}))
st.sidebar.markdown("**Credit Card Balance**")
new_cc_bal = st.sidebar.number_input(
    "Credit Card Owed", value=float(st.session_state.credit_card_balance),
    step=100.0, key="cc_balance_input", format="%.0f",
)
if new_cc_bal != st.session_state.credit_card_balance:
    st.session_state.credit_card_balance = new_cc_bal
    from tabs.shared import CREDIT_CARD_BAL_FILE
    CREDIT_CARD_BAL_FILE.write_text(json.dumps({"balance": new_cc_bal}))
st.sidebar.markdown("---")
sc1, sc2 = st.sidebar.columns(2)
if sc1.button("\U0001f4be Save All"):
    save_all()
    st.sidebar.success("All saved!")
if sc2.button("\U0001f4c2 Load All"):
    load_all()
    st.sidebar.success("All loaded!")
# -- Live prices + Net Worth --------------------------------------------------
st.sidebar.markdown("---")
all_tickers = get_all_tickers()
fetchable_tickers = tuple(sorted(
    t for t in all_tickers if t not in MONEY_MARKET_TICKERS and t not in MUTUAL_FUND_TICKERS
))
live_prices = fetch_live_prices(fetchable_tickers) if fetchable_tickers else {}
def live_net_worth():
    nw = -abs(st.session_state.car_loan_balance) - abs(st.session_state.credit_card_balance) + cash_balance
    for key, h in all_holdings.items():
        if not include_shared and h["account"] == "Shared":
            continue
        t = h["ticker"]
        nw += h["value"] if t in MONEY_MARKET_TICKERS else h["shares"] * live_prices.get(t, h["price"])
    return nw
live_nw = live_net_worth()
st.sidebar.metric("Live Net Worth", f"${live_nw:,.0f}")
if live_prices:
    st.sidebar.caption(f"Prices for {len(live_prices)}/{len(fetchable_tickers)} tickers")
else:
    st.sidebar.caption("Using position statement prices")
st.sidebar.caption(f"Growth: {growth_rate:.0%} | Sav Yield: {savings_yield:.2%}")
# -- Page routing -------------------------------------------------------------
if page == "12-Month Budget":
    from tabs.budget import render
    render(include_shared, net_per_check=NET_PER_CHECK, paydays_per_month=PAYDAYS_PER_MONTH)
elif page == "Portfolio Projection":
    from tabs.projection import render
    render(holdings_by_account, all_holdings, live_prices, cash_balance,
           growth_rate, savings_yield, include_shared, PORTFOLIO_NAMES)
elif page == "Analytics":
    from tabs.analytics import render
    render(holdings_by_account, all_holdings, live_prices, cash_balance,
           growth_rate, savings_yield, include_shared, PORTFOLIO_NAMES)
elif page == "Dividend Forecaster":
    from tabs.dividends import render
    render(include_shared, live_prices, growth_rate)
elif page == "Retirement & Paycheck":
    from tabs.retirement import render
    render(holdings_by_account, live_prices, cash_balance, growth_rate,
           savings_yield, include_shared)
elif page == "Balance Sheet":
    from tabs.balance_sheet import render
    render(holdings_by_account, all_holdings, live_prices, cash_balance, include_shared)
elif page == "Upload Positions":
    from tabs.upload import render
    render(parse_position_csv, build_holdings_from_positions, load_latest_positions,
           eq_df_loaded, ot_df_loaded, POSITION_DIR)
